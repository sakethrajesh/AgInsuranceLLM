import requests
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
import datetime

def evaluate(question):

    url = "http://localhost:5000/api/chat"
    data = {"messages": [{"role": "user", "content": question}]}
    headers = {'Content-Type': 'application/json'}
    response = requests.post(url, json=data, headers=headers)
    return response.json()["text"]


# Load the workbook
wb = load_workbook(filename='testbench.xlsx')
ws = wb.active

# Load the first column into a list
questions = [cell.value for cell in ws['A'] if cell.value is not None]

def main():
    for question in questions:
        answer = evaluate(question)
        
        # Find the last column with data
        last_column = ws.max_column
        
        # Get the current date
        current_date = datetime.date.today()
        
        # Write the current date to the first column of the next row
        ws.cell(row=ws.max_row+1, column=1, value=current_date)
        
        # Write the question and answer to the next available columns
        for idx, item in enumerate([question, answer], start=2):
            ws.cell(row=ws.max_row, column=idx, value=item)

    wb.save("questions.xlsx")
